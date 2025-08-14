import openpyxl
import re
import json

def parse_combinations_data(sheet):
    data = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0] or not str(row[0]).strip():
            continue

        rare_group_cell = str(row[0]).strip()
        
        match = re.search(r"RARE\[(\d+)\]", rare_group_cell)
        if not match:
            print(f"Skipping row with malformed RARE data: '{rare_group_cell}'")
            continue
            
        rare_group = int(match.group(1))

        combination = [
            int(row[1]) if row[1] and str(row[1]).strip() != '-' else None,
            int(row[2]) if row[2] and str(row[2]).strip() != '-' else None,
            int(row[3]) if row[3] and str(row[3]).strip() != '-' else None
        ]
        
        secondary_data = []
        if len(row) > 4 and row[4]:
            secondary_data_str = str(row[4]).strip()
            sub_matches = re.findall(r"\[([^\[\]]+)\]", secondary_data_str)
            for m in sub_matches:
                parts = [p.strip() for p in m.split(',')]
                secondary_data.append([int(p) if p.isdigit() else p for p in parts])

        if rare_group not in data:
            data[rare_group] = []
        
        data[rare_group].append({
            "combination": combination,
            "slots_info": secondary_data
        })
    return data

def parse_skills_data(sheet):
    data = {}
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if len(row) < 4 or not row[1]:
            continue
        
        try:
            group = int(row[1])
            skill_name = str(row[2]).strip()
            skill_level = int(row[3])

            if group not in data:
                data[group] = []
            
            data[group].append({
                "skill_name": skill_name,
                "skill_level": skill_level
            })
        except (ValueError, IndexError):
            print(f"Skipping row with invalid data: {row}")
            continue
            
    return data

def main():
    file_path = "data/amulettable.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found. Please ensure it's in the same directory as the script.")
        return

    try:
        combinations_sheet = workbook.worksheets[0]
        skills_sheet = workbook.worksheets[2]
    except IndexError:
        print("Error: The workbook does not contain enough sheets. Please ensure there is at least a first and third sheet.")
        return

    print("Parsing rarity combinations data...")
    combinations_data = parse_combinations_data(combinations_sheet)

    print("Parsing skills data...")
    skills_data = parse_skills_data(skills_sheet)

    final_data = {
        "rarity": combinations_data,
        "skills_data": skills_data
    }

    with open("src/extracted_data.json", "w", encoding="utf-8") as f:
        json.dump(final_data, f, indent=2, ensure_ascii=False)
    
    print("Data successfully extracted and saved to 'extracted_data.json'.")

if __name__ == "__main__":
    main()
